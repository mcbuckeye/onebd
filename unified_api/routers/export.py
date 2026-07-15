"""
Export endpoints for generating Excel and CSV files.
"""
from typing import Optional, List
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
import io
import csv
import structlog

logger = structlog.get_logger(__name__)

router = APIRouter()


class ExportRequest(BaseModel):
    """Request for deal export."""
    company: Optional[str] = None
    indication: Optional[str] = None
    technology: Optional[str] = None
    deal_type: Optional[str] = None
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    limit: int = 1000


@router.post("/export/deals/csv")
async def export_deals_csv(request: ExportRequest):
    """
    Export deals to CSV file.

    Returns a CSV file with deal data matching the specified filters.
    """
    from sqlalchemy import text
    from unified_api.services.database import get_cortellis_session

    logger.info("Exporting deals to CSV", filters=request.model_dump())

    # Build query with filters
    conditions = []
    params = {"limit": request.limit}

    if request.company:
        conditions.append("""
            d.id IN (
                SELECT dc.deal_id FROM deal_companies dc
                JOIN companies c ON c.id = dc.company_id
                WHERE c.name ILIKE :company
            )
        """)
        params["company"] = f"%{request.company}%"

    if request.indication:
        conditions.append("""
            d.id IN (
                SELECT di.deal_id FROM deal_indications di
                JOIN indications i ON i.id = di.indication_id
                WHERE i.name ILIKE :indication
            )
        """)
        params["indication"] = f"%{request.indication}%"

    if request.technology:
        conditions.append("""
            d.id IN (
                SELECT dt.deal_id FROM deal_technologies dt
                JOIN technologies t ON t.id = dt.technology_id
                WHERE t.name ILIKE :technology
            )
        """)
        params["technology"] = f"%{request.technology}%"

    if request.deal_type:
        conditions.append("d.deal_type ILIKE :deal_type")
        params["deal_type"] = f"%{request.deal_type}%"

    if request.date_from:
        conditions.append("d.date_start >= :date_from")
        params["date_from"] = request.date_from

    if request.date_to:
        conditions.append("d.date_start <= :date_to")
        params["date_to"] = request.date_to

    where_clause = " AND ".join(conditions) if conditions else "1=1"

    query = f"""
        SELECT
            d.id,
            d.title,
            d.deal_type,
            d.status,
            d.date_start::text,
            d.date_end::text,
            f.total_projected_current_amount as total_value,
            f.total_paid_amount as paid_value,
            (SELECT c.name FROM deal_companies dc
             JOIN companies c ON c.id = dc.company_id
             WHERE dc.deal_id = d.id AND dc.role = 'Principal' LIMIT 1) as principal_company,
            (SELECT c.name FROM deal_companies dc
             JOIN companies c ON c.id = dc.company_id
             WHERE dc.deal_id = d.id AND dc.role = 'Partner' LIMIT 1) as partner_company,
            (SELECT string_agg(i.name, '; ')
             FROM deal_indications di
             JOIN indications i ON i.id = di.indication_id
             WHERE di.deal_id = d.id) as indications,
            (SELECT string_agg(t.name, '; ')
             FROM deal_technologies dt
             JOIN technologies t ON t.id = dt.technology_id
             WHERE dt.deal_id = d.id) as technologies,
            (SELECT string_agg(dr.name_display, '; ')
             FROM deal_drugs dd
             JOIN drugs dr ON dr.id = dd.drug_id
             WHERE dd.deal_id = d.id) as drugs
        FROM deals d
        LEFT JOIN deal_finance_summary f ON f.deal_id = d.id
        WHERE {where_clause}
        ORDER BY d.date_start DESC NULLS LAST
        LIMIT :limit
    """

    with get_cortellis_session() as session:
        result = session.execute(text(query), params)
        rows = result.fetchall()

    # Generate CSV
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        'Deal ID', 'Title', 'Deal Type', 'Status',
        'Start Date', 'End Date', 'Total Value ($M)', 'Paid Value ($M)',
        'Principal Company', 'Partner Company',
        'Indications', 'Technologies', 'Drugs'
    ])

    # Data rows
    for row in rows:
        writer.writerow([
            row.id,
            row.title,
            row.deal_type,
            row.status,
            row.date_start,
            row.date_end,
            row.total_value,
            row.paid_value,
            row.principal_company,
            row.partner_company,
            row.indications,
            row.technologies,
            row.drugs,
        ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=deals_export.csv"}
    )


@router.get("/export/company/{company_id}/deals/csv")
async def export_company_deals_csv(
    company_id: int,
    limit: int = Query(1000, ge=1, le=10000),
):
    """
    Export all deals for a specific company to CSV.
    """
    from sqlalchemy import text
    from unified_api.services.database import get_cortellis_session

    logger.info("Exporting company deals to CSV", company_id=company_id)

    query = """
        SELECT
            d.id,
            d.title,
            d.deal_type,
            d.status,
            dc.role as company_role,
            d.date_start::text,
            d.date_end::text,
            f.total_projected_current_amount as total_value,
            (SELECT c2.name FROM deal_companies dc2
             JOIN companies c2 ON c2.id = dc2.company_id
             WHERE dc2.deal_id = d.id AND dc2.company_id != :company_id
             LIMIT 1) as counterparty,
            (SELECT string_agg(i.name, '; ')
             FROM deal_indications di
             JOIN indications i ON i.id = di.indication_id
             WHERE di.deal_id = d.id) as indications,
            (SELECT string_agg(dr.name_display, '; ')
             FROM deal_drugs dd
             JOIN drugs dr ON dr.id = dd.drug_id
             WHERE dd.deal_id = d.id) as drugs
        FROM deal_companies dc
        JOIN deals d ON d.id = dc.deal_id
        LEFT JOIN deal_finance_summary f ON f.deal_id = d.id
        WHERE dc.company_id = :company_id
        ORDER BY d.date_start DESC NULLS LAST
        LIMIT :limit
    """

    with get_cortellis_session() as session:
        # Get company name
        company_result = session.execute(text(
            "SELECT name FROM companies WHERE id = :company_id"
        ), {"company_id": company_id})
        company = company_result.fetchone()
        company_name = company.name if company else f"Company_{company_id}"

        result = session.execute(text(query), {"company_id": company_id, "limit": limit})
        rows = result.fetchall()

    # Generate CSV
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'Deal ID', 'Title', 'Deal Type', 'Status', 'Role',
        'Start Date', 'End Date', 'Total Value ($M)',
        'Counterparty', 'Indications', 'Drugs'
    ])

    for row in rows:
        writer.writerow([
            row.id,
            row.title,
            row.deal_type,
            row.status,
            row.company_role,
            row.date_start,
            row.date_end,
            row.total_value,
            row.counterparty,
            row.indications,
            row.drugs,
        ])

    output.seek(0)

    # Clean filename
    safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in company_name)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_deals.csv"}
    )


@router.get("/export/analytics/market-trends/csv")
async def export_market_trends_csv(
    granularity: str = Query("year", enum=["year", "quarter"]),
    therapy_area: Optional[str] = None,
    indication: Optional[str] = None,
    years: int = Query(10, ge=1, le=30),
):
    """
    Export market trends data to CSV.
    """
    from sqlalchemy import text
    from unified_api.services.database import get_cortellis_session

    logger.info("Exporting market trends to CSV")

    conditions = ["d.date_start IS NOT NULL"]
    params = {}

    if therapy_area:
        conditions.append("""
            d.id IN (
                SELECT di.deal_id FROM deal_indications di
                JOIN indications i ON i.id = di.indication_id
                WHERE i.name ILIKE :therapy_area
            )
        """)
        params["therapy_area"] = f"%{therapy_area}%"

    if indication:
        conditions.append("""
            d.id IN (
                SELECT di.deal_id FROM deal_indications di
                JOIN indications i ON i.id = di.indication_id
                WHERE i.name ILIKE :indication
            )
        """)
        params["indication"] = f"%{indication}%"

    where_clause = " AND ".join(conditions)

    if granularity == "quarter":
        period_expr = "EXTRACT(YEAR FROM d.date_start)::int || '-Q' || EXTRACT(QUARTER FROM d.date_start)::int"
        order_expr = "EXTRACT(YEAR FROM d.date_start), EXTRACT(QUARTER FROM d.date_start)"
    else:
        period_expr = "EXTRACT(YEAR FROM d.date_start)::int::text"
        order_expr = "EXTRACT(YEAR FROM d.date_start)"

    query = f"""
        SELECT
            {period_expr} as period,
            COUNT(*) as deal_count,
            SUM(f.total_projected_current_amount) as total_value,
            AVG(f.total_projected_current_amount) as avg_value,
            COUNT(f.total_projected_current_amount) as disclosed_count
        FROM deals d
        LEFT JOIN deal_finance_summary f ON f.deal_id = d.id
        WHERE {where_clause}
          AND d.date_start >= CURRENT_DATE - INTERVAL '{years} years'
        GROUP BY {order_expr}
        ORDER BY {order_expr} DESC
    """

    with get_cortellis_session() as session:
        result = session.execute(text(query), params)
        rows = result.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(['Period', 'Deal Count', 'Total Value ($M)', 'Average Value ($M)', 'Disclosed Deals'])

    for row in rows:
        writer.writerow([
            row.period,
            row.deal_count,
            round(row.total_value, 2) if row.total_value else '',
            round(row.avg_value, 2) if row.avg_value else '',
            row.disclosed_count,
        ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=market_trends.csv"}
    )


@router.get("/export/analytics/valuations/csv")
async def export_valuations_csv(
    benchmark_type: str = Query("phase", enum=["phase", "indication", "deal_type"]),
    years: int = Query(5, ge=1, le=20),
):
    """
    Export valuation benchmarks to CSV.
    """
    from sqlalchemy import text
    from unified_api.services.database import get_cortellis_session

    logger.info("Exporting valuations to CSV", benchmark_type=benchmark_type)

    if benchmark_type == "phase":
        query = f"""
            WITH deal_phases AS (
                SELECT DISTINCT ON (dte.deal_id)
                    dte.deal_id,
                    dte.stage
                FROM deal_timeline_events dte
                WHERE dte.stage IS NOT NULL AND dte.stage != ''
                ORDER BY dte.deal_id, dte.event_date ASC
            )
            SELECT
                dp.stage as category,
                COUNT(*) as deal_count,
                MIN(f.total_projected_current_amount) as min_value,
                MAX(f.total_projected_current_amount) as max_value,
                AVG(f.total_projected_current_amount) as avg_value,
                PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY f.total_projected_current_amount) as median_value,
                PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY f.total_projected_current_amount) as q1_value,
                PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY f.total_projected_current_amount) as q3_value
            FROM deal_phases dp
            JOIN deals d ON d.id = dp.deal_id
            JOIN deal_finance_summary f ON f.deal_id = d.id
            WHERE f.total_projected_current_amount IS NOT NULL
              AND d.date_start >= CURRENT_DATE - INTERVAL '{years} years'
            GROUP BY dp.stage
            ORDER BY deal_count DESC
        """
    elif benchmark_type == "indication":
        query = f"""
            SELECT
                i.name as category,
                COUNT(*) as deal_count,
                MIN(f.total_projected_current_amount) as min_value,
                MAX(f.total_projected_current_amount) as max_value,
                AVG(f.total_projected_current_amount) as avg_value,
                PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY f.total_projected_current_amount) as median_value,
                PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY f.total_projected_current_amount) as q1_value,
                PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY f.total_projected_current_amount) as q3_value
            FROM deal_indications di
            JOIN deals d ON d.id = di.deal_id
            JOIN indications i ON i.id = di.indication_id
            JOIN deal_finance_summary f ON f.deal_id = d.id
            WHERE f.total_projected_current_amount IS NOT NULL
              AND d.date_start >= CURRENT_DATE - INTERVAL '{years} years'
            GROUP BY i.name
            HAVING COUNT(*) >= 5
            ORDER BY deal_count DESC
        """
    else:  # deal_type
        query = f"""
            SELECT
                COALESCE(NULLIF(d.deal_type, ''), 'Unspecified') as category,
                COUNT(*) as deal_count,
                MIN(f.total_projected_current_amount) as min_value,
                MAX(f.total_projected_current_amount) as max_value,
                AVG(f.total_projected_current_amount) as avg_value,
                PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY f.total_projected_current_amount) as median_value,
                PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY f.total_projected_current_amount) as q1_value,
                PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY f.total_projected_current_amount) as q3_value
            FROM deals d
            JOIN deal_finance_summary f ON f.deal_id = d.id
            WHERE f.total_projected_current_amount IS NOT NULL
              AND d.date_start >= CURRENT_DATE - INTERVAL '{years} years'
            GROUP BY COALESCE(NULLIF(d.deal_type, ''), 'Unspecified')
            ORDER BY deal_count DESC
        """

    with get_cortellis_session() as session:
        result = session.execute(text(query))
        rows = result.fetchall()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'Category', 'Deal Count',
        'Min Value ($M)', 'Max Value ($M)', 'Average Value ($M)',
        'Median Value ($M)', '25th Percentile ($M)', '75th Percentile ($M)'
    ])

    for row in rows:
        writer.writerow([
            row.category,
            row.deal_count,
            round(row.min_value, 2) if row.min_value else '',
            round(row.max_value, 2) if row.max_value else '',
            round(row.avg_value, 2) if row.avg_value else '',
            round(row.median_value, 2) if row.median_value else '',
            round(row.q1_value, 2) if row.q1_value else '',
            round(row.q3_value, 2) if row.q3_value else '',
        ])

    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=valuations_by_{benchmark_type}.csv"}
    )


# ============================================
# Excel Export Endpoints
# ============================================

def _build_filter_conditions(request: ExportRequest) -> tuple:
    """Build SQL conditions from export request."""
    conditions = []
    params = {"limit": request.limit}

    if request.company:
        conditions.append("""
            d.id IN (
                SELECT dc.deal_id FROM deal_companies dc
                JOIN companies c ON c.id = dc.company_id
                WHERE c.name ILIKE :company
            )
        """)
        params["company"] = f"%{request.company}%"

    if request.indication:
        conditions.append("""
            d.id IN (
                SELECT di.deal_id FROM deal_indications di
                JOIN indications i ON i.id = di.indication_id
                WHERE i.name ILIKE :indication
            )
        """)
        params["indication"] = f"%{request.indication}%"

    if request.technology:
        conditions.append("""
            d.id IN (
                SELECT dt.deal_id FROM deal_technologies dt
                JOIN technologies t ON t.id = dt.technology_id
                WHERE t.name ILIKE :technology
            )
        """)
        params["technology"] = f"%{request.technology}%"

    if request.deal_type:
        conditions.append("d.agreement_type ILIKE :deal_type")
        params["deal_type"] = f"%{request.deal_type}%"

    if request.date_from:
        conditions.append("d.date_start >= :date_from")
        params["date_from"] = request.date_from

    if request.date_to:
        conditions.append("d.date_start <= :date_to")
        params["date_to"] = request.date_to

    return conditions, params


@router.post("/export/deals/excel")
async def export_deals_excel(request: ExportRequest):
    """
    Export deals to Excel file (.xlsx).

    Returns an Excel file with deal data matching the specified filters.
    Includes formatted headers and auto-sized columns.
    """
    from sqlalchemy import text
    from unified_api.services.database import get_cortellis_session
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    logger.info("Exporting deals to Excel", filters=request.model_dump())

    conditions, params = _build_filter_conditions(request)
    where_clause = " AND ".join(conditions) if conditions else "1=1"

    query = f"""
        SELECT
            d.id,
            d.title,
            d.agreement_type,
            d.deal_type,
            d.status,
            d.date_start::text,
            d.date_end::text,
            ta.name as therapy_area,
            d.phase_highest_start,
            f.total_projected_current_amount as total_value,
            f.total_paid_amount as paid_value,
            (SELECT c.name FROM deal_companies dc
             JOIN companies c ON c.id = dc.company_id
             WHERE dc.deal_id = d.id AND dc.role = 'Principal' LIMIT 1) as principal_company,
            (SELECT c.name FROM deal_companies dc
             JOIN companies c ON c.id = dc.company_id
             WHERE dc.deal_id = d.id AND dc.role = 'Partner' LIMIT 1) as partner_company,
            (SELECT string_agg(i.name, '; ')
             FROM deal_indications di
             JOIN indications i ON i.id = di.indication_id
             WHERE di.deal_id = d.id) as indications,
            (SELECT string_agg(t.name, '; ')
             FROM deal_technologies dt
             JOIN technologies t ON t.id = dt.technology_id
             WHERE dt.deal_id = d.id) as technologies,
            (SELECT string_agg(dr.name_display, '; ')
             FROM deal_drugs dd
             JOIN drugs dr ON dr.id = dd.drug_id
             WHERE dd.deal_id = d.id) as drugs
        FROM deals d
        LEFT JOIN deal_finance_summary f ON f.deal_id = d.id
        LEFT JOIN therapy_areas ta ON ta.id = d.therapy_area_id
        WHERE {where_clause}
        ORDER BY d.date_start DESC NULLS LAST
        LIMIT :limit
    """

    with get_cortellis_session() as session:
        result = session.execute(text(query), params)
        rows = result.fetchall()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Deals"

    # Headers
    headers = [
        'Deal ID', 'Title', 'Agreement Type', 'Deal Type', 'Status',
        'Start Date', 'End Date', 'Therapy Area', 'Phase at Start',
        'Total Value ($M)', 'Paid Value ($M)',
        'Principal Company', 'Partner Company',
        'Indications', 'Technologies', 'Drugs'
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.id)
        ws.cell(row=row_idx, column=2, value=row.title)
        ws.cell(row=row_idx, column=3, value=row.agreement_type)
        ws.cell(row=row_idx, column=4, value=row.deal_type)
        ws.cell(row=row_idx, column=5, value=row.status)
        ws.cell(row=row_idx, column=6, value=row.date_start)
        ws.cell(row=row_idx, column=7, value=row.date_end)
        ws.cell(row=row_idx, column=8, value=row.therapy_area)
        ws.cell(row=row_idx, column=9, value=row.phase_highest_start)
        ws.cell(row=row_idx, column=10, value=row.total_value)
        ws.cell(row=row_idx, column=11, value=row.paid_value)
        ws.cell(row=row_idx, column=12, value=row.principal_company)
        ws.cell(row=row_idx, column=13, value=row.partner_company)
        ws.cell(row=row_idx, column=14, value=row.indications)
        ws.cell(row=row_idx, column=15, value=row.technologies)
        ws.cell(row=row_idx, column=16, value=row.drugs)

    # Auto-size columns (approximate)
    column_widths = [10, 50, 20, 15, 12, 12, 12, 20, 15, 15, 15, 30, 30, 40, 30, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deals_export.xlsx"}
    )


@router.get("/export/company/{company_id}/deals/excel")
async def export_company_deals_excel(
    company_id: int,
    limit: int = Query(1000, ge=1, le=10000),
):
    """
    Export all deals for a specific company to Excel.
    """
    from sqlalchemy import text
    from unified_api.services.database import get_cortellis_session
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    logger.info("Exporting company deals to Excel", company_id=company_id)

    query = """
        SELECT
            d.id,
            d.title,
            d.agreement_type,
            d.deal_type,
            d.status,
            dc.role as company_role,
            d.date_start::text,
            d.date_end::text,
            ta.name as therapy_area,
            f.total_projected_current_amount as total_value,
            (SELECT c2.name FROM deal_companies dc2
             JOIN companies c2 ON c2.id = dc2.company_id
             WHERE dc2.deal_id = d.id AND dc2.company_id != :company_id
             LIMIT 1) as counterparty,
            (SELECT string_agg(i.name, '; ')
             FROM deal_indications di
             JOIN indications i ON i.id = di.indication_id
             WHERE di.deal_id = d.id) as indications,
            (SELECT string_agg(dr.name_display, '; ')
             FROM deal_drugs dd
             JOIN drugs dr ON dr.id = dd.drug_id
             WHERE dd.deal_id = d.id) as drugs
        FROM deal_companies dc
        JOIN deals d ON d.id = dc.deal_id
        LEFT JOIN deal_finance_summary f ON f.deal_id = d.id
        LEFT JOIN therapy_areas ta ON ta.id = d.therapy_area_id
        WHERE dc.company_id = :company_id
        ORDER BY d.date_start DESC NULLS LAST
        LIMIT :limit
    """

    with get_cortellis_session() as session:
        # Get company name
        company_result = session.execute(text(
            "SELECT name FROM companies WHERE id = :company_id"
        ), {"company_id": company_id})
        company = company_result.fetchone()
        company_name = company.name if company else f"Company_{company_id}"

        result = session.execute(text(query), {"company_id": company_id, "limit": limit})
        rows = result.fetchall()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Deals"

    # Headers
    headers = [
        'Deal ID', 'Title', 'Agreement Type', 'Deal Type', 'Status', 'Role',
        'Start Date', 'End Date', 'Therapy Area', 'Total Value ($M)',
        'Counterparty', 'Indications', 'Drugs'
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.id)
        ws.cell(row=row_idx, column=2, value=row.title)
        ws.cell(row=row_idx, column=3, value=row.agreement_type)
        ws.cell(row=row_idx, column=4, value=row.deal_type)
        ws.cell(row=row_idx, column=5, value=row.status)
        ws.cell(row=row_idx, column=6, value=row.company_role)
        ws.cell(row=row_idx, column=7, value=row.date_start)
        ws.cell(row=row_idx, column=8, value=row.date_end)
        ws.cell(row=row_idx, column=9, value=row.therapy_area)
        ws.cell(row=row_idx, column=10, value=row.total_value)
        ws.cell(row=row_idx, column=11, value=row.counterparty)
        ws.cell(row=row_idx, column=12, value=row.indications)
        ws.cell(row=row_idx, column=13, value=row.drugs)

    # Auto-size columns
    column_widths = [10, 50, 20, 15, 12, 12, 12, 12, 20, 15, 30, 40, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Clean filename
    safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in company_name)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe_name}_deals.xlsx"}
    )


@router.get("/export/search-results/excel")
async def export_search_results_excel(
    company: Optional[str] = None,
    therapy_area: Optional[str] = None,
    indication: Optional[List[str]] = Query(None),
    technology: Optional[List[str]] = Query(None),
    deal_type: Optional[List[str]] = Query(None),
    phase: Optional[List[str]] = Query(None),
    status: Optional[List[str]] = Query(None),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    value_min: Optional[float] = None,
    value_max: Optional[float] = None,
    disclosed_only: bool = False,
    limit: int = Query(1000, ge=1, le=10000),
):
    """
    Export search results to Excel with advanced filtering.

    Supports the same filters as the advanced search endpoint.
    """
    from sqlalchemy import text
    from unified_api.services.database import get_cortellis_session
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    logger.info("Exporting search results to Excel")

    # Build query with filters
    conditions = []
    params = {"limit": limit}
    joins = []

    if company:
        conditions.append("""
            d.id IN (
                SELECT dc.deal_id FROM deal_companies dc
                JOIN companies c ON c.id = dc.company_id
                WHERE c.name ILIKE :company
            )
        """)
        params["company"] = f"%{company}%"

    if deal_type:
        conditions.append("d.agreement_type = ANY(:deal_types)")
        params["deal_types"] = deal_type

    if date_from:
        conditions.append("d.date_start >= :date_from")
        params["date_from"] = date_from

    if date_to:
        conditions.append("d.date_start <= :date_to")
        params["date_to"] = date_to

    if therapy_area:
        joins.append("LEFT JOIN therapy_areas ta ON ta.id = d.therapy_area_id")
        conditions.append("ta.name ILIKE :therapy_area")
        params["therapy_area"] = f"%{therapy_area}%"

    if indication:
        conditions.append("""
            d.id IN (
                SELECT di.deal_id FROM deal_indications di
                JOIN indications i ON i.id = di.indication_id
                WHERE i.name ILIKE ANY(:indications)
            )
        """)
        params["indications"] = [f"%{ind}%" for ind in indication]

    if technology:
        conditions.append("""
            d.id IN (
                SELECT dt.deal_id FROM deal_technologies dt
                JOIN technologies t ON t.id = dt.technology_id
                WHERE t.name ILIKE ANY(:technologies)
            )
        """)
        params["technologies"] = [f"%{tech}%" for tech in technology]

    if phase:
        conditions.append("d.phase_highest_start = ANY(:phases)")
        params["phases"] = phase

    if value_min is not None:
        conditions.append("f.total_projected_current_amount >= :value_min")
        params["value_min"] = value_min

    if value_max is not None:
        conditions.append("f.total_projected_current_amount <= :value_max")
        params["value_max"] = value_max

    if disclosed_only:
        conditions.append("f.total_projected_current_amount IS NOT NULL")

    if status:
        conditions.append("d.status = ANY(:statuses)")
        params["statuses"] = status

    where_clause = " AND ".join(conditions) if conditions else "1=1"
    join_clause = " ".join(joins)

    # Add therapy_areas join if not already present
    if "therapy_areas ta" not in join_clause:
        join_clause = "LEFT JOIN therapy_areas ta ON ta.id = d.therapy_area_id " + join_clause

    query = f"""
        SELECT
            d.id,
            d.title,
            d.agreement_type,
            d.deal_type,
            d.status,
            d.date_start::text,
            d.date_end::text,
            ta.name as therapy_area,
            d.phase_highest_start,
            f.total_projected_current_amount as total_value,
            f.total_paid_amount as paid_value,
            (SELECT c.name FROM deal_companies dc
             JOIN companies c ON c.id = dc.company_id
             WHERE dc.deal_id = d.id AND dc.role = 'Principal' LIMIT 1) as principal_company,
            (SELECT c.name FROM deal_companies dc
             JOIN companies c ON c.id = dc.company_id
             WHERE dc.deal_id = d.id AND dc.role = 'Partner' LIMIT 1) as partner_company,
            (SELECT string_agg(i.name, '; ')
             FROM deal_indications di
             JOIN indications i ON i.id = di.indication_id
             WHERE di.deal_id = d.id) as indications,
            (SELECT string_agg(t.name, '; ')
             FROM deal_technologies dt
             JOIN technologies t ON t.id = dt.technology_id
             WHERE dt.deal_id = d.id) as technologies,
            (SELECT string_agg(dr.name_display, '; ')
             FROM deal_drugs dd
             JOIN drugs dr ON dr.id = dd.drug_id
             WHERE dd.deal_id = d.id) as drugs
        FROM deals d
        LEFT JOIN deal_finance_summary f ON f.deal_id = d.id
        {join_clause}
        WHERE {where_clause}
        ORDER BY d.date_start DESC NULLS LAST
        LIMIT :limit
    """

    with get_cortellis_session() as session:
        result = session.execute(text(query), params)
        rows = result.fetchall()

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"

    # Headers
    headers = [
        'Deal ID', 'Title', 'Agreement Type', 'Deal Type', 'Status',
        'Start Date', 'End Date', 'Therapy Area', 'Phase at Start',
        'Total Value ($M)', 'Paid Value ($M)',
        'Principal Company', 'Partner Company',
        'Indications', 'Technologies', 'Drugs'
    ]

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.id)
        ws.cell(row=row_idx, column=2, value=row.title)
        ws.cell(row=row_idx, column=3, value=row.agreement_type)
        ws.cell(row=row_idx, column=4, value=row.deal_type)
        ws.cell(row=row_idx, column=5, value=row.status)
        ws.cell(row=row_idx, column=6, value=row.date_start)
        ws.cell(row=row_idx, column=7, value=row.date_end)
        ws.cell(row=row_idx, column=8, value=row.therapy_area)
        ws.cell(row=row_idx, column=9, value=row.phase_highest_start)
        ws.cell(row=row_idx, column=10, value=row.total_value)
        ws.cell(row=row_idx, column=11, value=row.paid_value)
        ws.cell(row=row_idx, column=12, value=row.principal_company)
        ws.cell(row=row_idx, column=13, value=row.partner_company)
        ws.cell(row=row_idx, column=14, value=row.indications)
        ws.cell(row=row_idx, column=15, value=row.technologies)
        ws.cell(row=row_idx, column=16, value=row.drugs)

    # Auto-size columns
    column_widths = [10, 50, 20, 15, 12, 12, 12, 20, 15, 15, 15, 30, 30, 40, 30, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=search_results.xlsx"}
    )
